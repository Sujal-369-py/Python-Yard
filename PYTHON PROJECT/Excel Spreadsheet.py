import openpyxl as xl

# Load the workbook
wb = xl.load_workbook("transactions(unchanged).xlsx")

# Make sure the sheet name is correct
sheet = wb['Sheet1']  # Correct the sheet name here

# Access the cell in the sheet correctly
cell = sheet.cell(1, 1)

# Print the value of the cell
# print(cell.value)
# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

wb.save("newtransaction.xlsx")

